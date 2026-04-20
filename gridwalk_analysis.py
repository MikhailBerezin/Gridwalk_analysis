"""
Gridwalk foot-fault analysis for nerve window paper (Figure 1h).

Loads gridwalk scoring data from an Excel workbook, computes per-bin mean
+/- SEM of the weighted success rate for the windowed (right) and
contralateral (left) hindlimb, and produces the Figure 1h bar plot.

Two scoring conventions are present in the data and both are handled:

    Letter-based (older rows) in columns K/L/M/N:
        K = count of "G" (good step)  --> score 1.0
        L = count of "M" (medium)      --> score 0.5
        M = count of "P" (poor)        --> score 0.0
        N = count of "B" (bad/missed)  --> score 0.0
        Weighted success = (K + 0.5*L) / (K + L + M + N)

    Digit-based (newer rows) in columns O/P/Q/R/S:
        O = count of 0 (complete miss)          --> score 0.00
        P = count of 1                          --> score 0.25
        Q = count of 2                          --> score 0.50
        R = count of 3                          --> score 0.75
        S = count of 4 (clean step)             --> score 1.00
        Weighted success = (S + 0.75*R + 0.5*Q + 0.25*P) / sum(O..S)

Required Excel columns:
    A: Mouse identifier       C: Foot L/R         E: Days after implant
    K-N: letter counts (optional)  O-S: digit counts (optional)

Usage: python gridwalk_analysis.py path/to/data.xlsx

Dependencies: openpyxl, numpy, matplotlib.
"""

import argparse
import numpy as np
import matplotlib.pyplot as plt
import openpyxl


BINS = [(-9, 0), (1, 7), (8, 13), (14, 25), (26, 35)]


def success_rate(row_counts):
    """Compute weighted success rate for one row. Returns None if no counts."""
    K, L, Mc, N, O, P, Q, R, S = (x or 0 for x in row_counts)
    letter_total = K + L + Mc + N
    digit_total = O + P + Q + R + S
    if digit_total > 0:
        return (S + 0.75 * R + 0.5 * Q + 0.25 * P) / digit_total
    if letter_total > 0:
        return (K + 0.5 * L) / letter_total
    return None


def load(xlsx_path):
    """Return (days, success, side) arrays."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    days, success, side = [], [], []
    for row in range(2, ws.max_row + 1):
        d = ws.cell(row, 5).value
        s = ws.cell(row, 3).value
        if d is None or s is None:
            continue
        counts = tuple(ws.cell(row, c).value for c in range(11, 20))  # K..S
        T = success_rate(counts)
        if T is None:
            continue
        side_s = str(s).strip().lower()
        if side_s in ("l", "left"):
            side.append("L")
        elif side_s in ("r", "right"):
            side.append("R")
        else:
            continue
        days.append(float(d))
        success.append(T)
    return np.array(days), np.array(success), np.array(side)


def bin_stats(days, rates):
    means, sems = [], []
    for b_start, b_end in BINS:
        data = rates[(days >= b_start) & (days <= b_end)]
        if len(data):
            means.append(data.mean() * 100)
            sems.append((data.std(ddof=1) / np.sqrt(len(data))) * 100)
        else:
            means.append(np.nan)
            sems.append(np.nan)
    return np.array(means), np.array(sems)


def plot_fig1h(xlsx_path, outpath=None):
    days, success, side = load(xlsx_path)
    L = side == "L"
    R = side == "R"

    means_L, sems_L = bin_stats(days[L], success[L])
    means_R, sems_R = bin_stats(days[R], success[R])

    labels = []
    for i, (b_start, b_end) in enumerate(BINS):
        if i == 0:
            labels.append("<0")
        else:
            in_bin = days[(days >= b_start) & (days <= b_end)]
            labels.append(f"{int(in_bin.min())}-{int(in_bin.max())}" if len(in_bin) else f"{b_start}-{b_end}")

    # Report numbers
    print(f"{'Bin':<8} {'L mean±SEM':<20} {'R mean±SEM':<20}")
    for i, lbl in enumerate(labels):
        print(f"{lbl:<8} {means_L[i]:.2f} +/- {sems_L[i]:.2f}      {means_R[i]:.2f} +/- {sems_R[i]:.2f}")

    x = np.arange(len(BINS))
    bw = 0.35
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(x - bw / 2, means_L, width=bw, yerr=sems_L, capsize=5,
           color="blue", alpha=0.8, label="Contralateral leg")
    ax.bar(x + bw / 2, means_R, width=bw, yerr=sems_R, capsize=5,
           color="red", alpha=0.8, label="Windowed leg")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=16)
    ax.set_xlabel("Time after implantation (days)", fontsize=18)
    ax.set_ylabel("Weighted Success Rate (%)", fontsize=18)
    ax.set_ylim(0, max(100, np.ceil(max(np.nanmax(means_L), np.nanmax(means_R)) / 10) * 10))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(loc="lower right", fontsize=14)
    plt.tight_layout()
    if outpath:
        plt.savefig(outpath, dpi=200)
    plt.show()


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("xlsx", help="Path to gridwalk Excel workbook")
    p.add_argument("--out", default="figure_1h.png", help="Output PNG path")
    args = p.parse_args()
    plot_fig1h(args.xlsx, outpath=args.out)
